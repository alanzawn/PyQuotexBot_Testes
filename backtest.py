import pandas as pd
import asyncio
import time

from pyquotex.quotexapi.config import email, password
from pyquotex.quotexapi.stable_api import Quotex
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side


# Função para calcular as bandas (exemplo simples de bandas de Bollinger)
def calculate_bands(df, window=20, num_std=2):
    # Cálculo da média móvel (simples)
    df['sma'] = df['close'].rolling(window=window).mean()
    # Cálculo do desvio padrão
    df['std'] = df['close'].rolling(window=window).std()
    # Cálculo das bandas
    df['upperBand'] = df['sma'] + (df['std'] * num_std)
    df['lowerBand'] = df['sma'] - (df['std'] * num_std)
    return df


# Função de backtest
# Função de backtest
def backtest(df, asset_name):
    # Lista para armazenar os resultados do backtest
    results = []

    # Iterar através das velas a partir da quinta (i+1) até a penúltima (i+4) para garantir a análise das velas anteriores
    for i in range(4, len(df) - 5):  # -1 para garantir que não acesse um índice fora do limite
        # Pegando as velas de análise e as 3 velas anteriores para os tick volumes
        candle_analysis = df.iloc[i]
        prev_candles = df.iloc[i - 4:i]

        tick_volumes = prev_candles['ticks'].values  # Últimos 3 ticks (volumes)

        # Condições para a compra
        if (candle_analysis['low'] <= candle_analysis['lowerBand'] and
                candle_analysis['close'] > candle_analysis['open'] and
                candle_analysis['close'] > candle_analysis['lowerBand'] and
                candle_analysis['ticks'] > tick_volumes[:3].mean()):
            # Se atender à condição de compra
            buy_price = candle_analysis['low'] - 0.00005  # Adicionar um pequeno desconto
            entry_type = "Compra"
            next_candle = df.iloc[i + 4]  # Agora estamos seguros que i + 1 não excederá o limite

            # Verifica o resultado (WIN ou LOSS) baseado na vela seguinte
            result = "win" if next_candle['close'] > buy_price else "loss"

            results.append({
                'Time': None,
                'Ativo': asset_name,
                'Operação': entry_type,
                'Entrada': buy_price,
                'Tick': candle_analysis['ticks'],
                'Média Ticks': tick_volumes[:3].mean(),
                'Resultado': result
            })

        # Condições para a venda
        elif (candle_analysis['high'] >= candle_analysis['upperBand'] and
              candle_analysis['close'] < candle_analysis['open'] and
              candle_analysis['close'] < candle_analysis['upperBand'] and
              candle_analysis['ticks'] > tick_volumes[:3].mean()):
            # Se atender à condição de venda
            sell_price = candle_analysis['high'] + 0.00005  # Adicionar um pequeno acréscimo
            entry_type = "Venda"
            next_candle = df.iloc[i + 4]  # Agora estamos seguros que i + 1 não excederá o limite

            # Verifica o resultado (WIN ou LOSS) baseado na vela seguinte
            result = "win" if next_candle['close'] < sell_price else "loss"

            results.append({
                'Time': None,
                'Ativo': asset_name,
                'Operação': entry_type,
                'Entrada': sell_price,
                'Tick': candle_analysis['ticks'],
                'Média Ticks': tick_volumes[:3].mean(),
                'Resultado': result
            })

    # Criando um DataFrame com os resultados
    results_df = pd.DataFrame(results)
    return results_df



# Função para gerar o arquivo Excel formatado
def save_to_excel(df, filename='backtest_results.xlsx'):
    # Criar um arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Backtest"

    # Definir um estilo para os cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    # Escrever os cabeçalhos
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border

    # Definir um estilo para as células de dados
    data_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    # Preencher as células com os dados e aplicar borda
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_data)
            cell.border = data_border

    # Ajustar a largura das colunas automaticamente
    for col in range(1, len(df.columns) + 1):
        max_length = 0
        column = chr(64 + col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Salvar o arquivo Excel
    wb.save(filename)


# Função para pegar os dados de todos os ativos e realizar o backtest
async def get_candles_all_asset(client):
    start_time_total = time.time()  # Marca o tempo total de execução

    check_connect, message = await client.connect()
    if check_connect:
        offset = 3600  # in seconds
        period = 60  # in seconds
        codes_asset = await client.get_all_assets()

        all_results = []  # Lista para armazenar os resultados de todos os ativos

        # Loop para processar cada ativo
        for asset in codes_asset.keys():
            asset_name, asset_data = await client.get_available_asset(asset)
            if asset_data[2] and '/' in asset_data[1] and not 'BRL' in asset_data[1] and 'OTC' in asset_data[1]:
                print(f"Processando o ativo: {asset_name}...")

                # Coletar os candles
                end_from_time = time.time()
                candles = await client.get_candles(asset, end_from_time, offset, period)

                rates_frame = pd.DataFrame(candles)

                # Calcular as bandas e adicionar aos dados
                rates_frame = calculate_bands(rates_frame)

                # Inicia o contador de tempo para o backtest do ativo
                start_time_asset = time.time()

                # Executa o backtest
                backtest_results_df = backtest(rates_frame, asset_name)

                # Adiciona os resultados do ativo ao total
                all_results.append(backtest_results_df)

                # Calcula o tempo de execução do backtest para o ativo
                end_time_asset = time.time()
                elapsed_time_asset = end_time_asset - start_time_asset
                print(f"Backtest do ativo {asset_name} concluído em {elapsed_time_asset:.2f} segundos.")

                #await asyncio.sleep(1)  # Pausa entre as requisições

        # Juntando todos os resultados
        final_results_df = pd.concat(all_results, ignore_index=True)

        # Calcula o tempo total de execução
        end_time_total = time.time()
        elapsed_time_total = end_time_total - start_time_total
        print(f"Backtest de todos os ativos concluído em {elapsed_time_total:.2f} segundos.")

        # Exibe os resultados finais e salva em CSV e XLSX
        print(final_results_df)
        final_results_df.to_csv('final_backtest_results.csv', index=False)
        save_to_excel(final_results_df, 'final_backtest_results.xlsx')

    client.close()


# Função principal
async def main():
    client = Quotex(email=email, password=password)

    # Executa a função de obter os candles de todos os ativos
    await get_candles_all_asset(client)


# Executa a função assíncrona
asyncio.run(main())
